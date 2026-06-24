#!/usr/bin/env python3
"""Ingest approved NBA Excel sheets into normalized SQLite and JSON outputs."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import sqlite3
import sys
import unicodedata
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "raw" / "nba_data_2025_26.xlsx"
DEFAULT_SQLITE = ROOT / "data" / "processed" / "nba_master.sqlite"
DEFAULT_COLUMN_DICTIONARY = ROOT / "data" / "processed" / "column_dictionary.json"
DEFAULT_ISSUES_LOG = ROOT / "data" / "processed" / "data_issues_log.json"
DEFAULT_PLAYERS_JSON = ROOT / "public" / "data" / "players.json"
DEFAULT_PROFILE_DIR = ROOT / "public" / "data" / "player_profiles"
DEFAULT_RUNTIME_SUMMARIES = ROOT / "src" / "lib" / "data" / "generated" / "master-player-summaries.json"
DEFAULT_OFFICIAL_SNAPSHOT = ROOT / "src" / "lib" / "data" / "generated" / "official-snapshot.json"

SEASON = "2025-26"
SEASON_TYPE = "Regular Season"
PROFILE_STAT_LIMIT = None
MISSING_MARKERS = {"", "-", "--", "—", "n/a", "na", "null"}
PLAYER_HEADER_NAMES = {"player", "vs_player"}
TEAM_HEADER_NAMES = {"team"}
REVIEWED_IMPORT_SHEETS = {
    "General - Official Leaders": "Reviewed after audit: NBA.com official leaders export with a time-formatted 3PM header.",
    "General - Traditional": "Reviewed after audit: NBA.com Per Game export with one time-formatted 3PM header.",
    "General - Estimated Advanced": "Reviewed after audit: NBA.com estimated advanced export with player rows but no team column.",
    "Clutch - Traditional": "Reviewed after audit: NBA.com clutch traditional export with one time-formatted 3PM header.",
    "Sheet69": "Reviewed after audit: generic sheet name, but layout is a valid NBA.com defensive play type export.",
    "Sheet70": "Reviewed after audit: generic sheet name, but layout is a valid NBA.com defensive play type export.",
    "Tracking - Catch & Shoot": "Reviewed after audit: NBA.com tracking export with one time-formatted 3PM header.",
    "Tracking - Pullup Shooting": "Reviewed after audit: NBA.com tracking export with one time-formatted 3PM header.",
    "Shooting Dashboard - Overall": "Reviewed after audit: NBA.com grouped shooting dashboard with a time-formatted 3FGM header.",
    "Shooting Dashboard - Catch & Sh": "Reviewed after audit: NBA.com grouped shooting dashboard with a time-formatted 3FGM header.",
    "Shooting Dashboard - Pullups": "Reviewed after audit: NBA.com grouped shooting dashboard with a time-formatted 3FGM header.",
    "Shooting - 5ft Range": "Reviewed after audit: NBA.com grouped shooting zone export; group labels are flattened into stat names.",
    "Shooting - 8ft Range": "Reviewed after audit: NBA.com grouped shooting zone export; group labels are flattened into stat names.",
    "Shooting - By Zone": "Reviewed after audit: NBA.com grouped shooting zone export; group labels are flattened into stat names.",
    "Opponent Shooting - 5ft Range": "Reviewed after audit: NBA.com grouped opponent shooting zone export; group labels are flattened into stat names.",
    "Opponent Shooting - 8ft Range": "Reviewed after audit: NBA.com grouped opponent shooting zone export; group labels are flattened into stat names.",
    "Opponent Shooting - By Zone": "Reviewed after audit: NBA.com grouped opponent shooting zone export; group labels are flattened into stat names.",
    "Bios": "Reviewed after audit: NBA.com bio export with height cells sometimes formatted as Excel dates.",
}
REVIEWED_HEADER_REPLACEMENTS = {
    "General - Traditional": {
        1: "Rank",
        13: "3PM",
        30: "+/-",
    }
}
GROUPED_HEADER_SHEETS = {
    "Shooting Dashboard - Overall",
    "Shooting Dashboard - Catch & Sh",
    "Shooting Dashboard - Pullups",
    "Shooting - 5ft Range",
    "Shooting - 8ft Range",
    "Shooting - By Zone",
    "Opponent Shooting - 5ft Range",
    "Opponent Shooting - 8ft Range",
    "Opponent Shooting - By Zone",
}
RUNTIME_SUMMARY_SHEETS = {
    "General - Traditional",
    "General - Advanced",
    "Bios",
    "General - Defense",
    "General - Misc",
    "General - Usage",
}


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def excel_column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.replace(".", "")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    return slug or "unknown-player"


def canonical_player_name(raw_name: Any) -> tuple[str, str | None]:
    name = re.sub(r"\s+", " ", text_value(raw_name)).strip()
    if not name:
        return "", None
    if "," not in name:
        return name, None
    last, first = [part.strip() for part in name.split(",", 1)]
    if not last or not first:
        return name, None
    return f"{first} {last}".strip(), "last_first_name_reordered"


def snake_case(value: Any) -> str:
    text = text_value(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.replace("\n", " ").replace("\r", " ").strip().lower()
    replacements = [
        (r"\b3pm\b", "three_pm"),
        (r"\b3pa\b", "three_pa"),
        (r"\b3pt\b", "three_pt"),
        (r"\b3p\b", "three_p"),
        (r"\b3fgm\b", "three_fgm"),
        (r"\b2pt\b", "two_pt"),
        (r"\b2fgm\b", "two_fgm"),
        (r"\b2nd\b", "second"),
        (r"\boff\b", "off"),
        (r"\bdef\b", "def"),
    ]
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text)
    text = text.replace("%", " pct ")
    text = text.replace("#", " number ")
    text = text.replace("+", " plus ")
    text = text.replace("<", " lt ")
    text = text.replace(">", " gt ")
    text = text.replace("/", " per ")
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        return "unnamed_column"
    if text[0].isdigit():
        text = f"stat_{text}"
    return text


def height_inches_from_value(value: Any) -> tuple[float | None, str | None]:
    if isinstance(value, datetime):
        feet = value.month
        inches = value.day
        if 4 <= feet <= 8 and 0 <= inches <= 11:
            return float(feet * 12 + inches), "height_excel_date_converted_to_inches"
    if isinstance(value, date):
        feet = value.month
        inches = value.day
        if 4 <= feet <= 8 and 0 <= inches <= 11:
            return float(feet * 12 + inches), "height_excel_date_converted_to_inches"
    text = text_value(value).strip()
    match = re.match(r"^([4-8])-(\d{1,2})$", text)
    if match:
        feet = int(match.group(1))
        inches = int(match.group(2))
        if 0 <= inches <= 11:
            return float(feet * 12 + inches), "height_text_converted_to_inches"
    return None, None


def clean_numeric(value: Any, *, source_sheet: str = "", cleaned_column_name: str = "") -> tuple[float | None, str | None]:
    if value is None:
        return None, "blank_value"
    if source_sheet == "Bios" and cleaned_column_name == "height":
        height_inches, note = height_inches_from_value(value)
        if height_inches is not None:
            return height_inches, note
    if isinstance(value, bool):
        return None, "boolean_value_not_numeric"
    if isinstance(value, (datetime, date, time)):
        return None, "date_or_time_value_not_numeric"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None, "non_finite_numeric_value"
        return float(value), None
    text = text_value(value).strip()
    if text.lower() in MISSING_MARKERS:
        return None, "blank_value"
    cleaned = text.replace(",", "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1].strip()
    cleaned = cleaned.replace("$", "").strip()
    try:
        return float(cleaned), None
    except ValueError:
        return None, "non_numeric_text_value"


def add_issue(
    issues: list[dict[str, Any]],
    severity: str,
    issue_type: str,
    message: str,
    **details: Any,
) -> None:
    issues.append(
        {
            "severity": severity,
            "type": issue_type,
            "message": message,
            "details": details,
        }
    )


def load_sheet_decisions(audit: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {sheet["sheetName"]: sheet for sheet in audit["sheets"]}


def find_role_columns(headers: list[Any], cleaned_overrides: dict[int, str] | None = None) -> tuple[int | None, int | None]:
    player_col = None
    team_col = None
    for idx, header in enumerate(headers, start=1):
        normalized = cleaned_overrides.get(idx, snake_case(header)) if cleaned_overrides else snake_case(header)
        if player_col is None and normalized in PLAYER_HEADER_NAMES:
            player_col = idx
        if team_col is None and normalized in TEAM_HEADER_NAMES:
            team_col = idx
    return player_col, team_col


def is_time_formatted_three_header(value: Any) -> bool:
    if isinstance(value, time):
        return value.hour == 15 and value.minute == 0 and value.second == 0
    if isinstance(value, datetime):
        return value.hour == 15 and value.minute == 0 and value.second == 0
    return text_value(value).strip().lower() in {"15:00:00", "15:00"}


def row_values(ws: Any, row_number: int) -> list[Any]:
    return [cell.value for cell in next(ws.iter_rows(min_row=row_number, max_row=row_number))]


def infer_header_row(ws: Any, sheet_name: str, audit_header_row: int | None, issues: list[dict[str, Any]]) -> int | None:
    candidates: list[tuple[int, int, bool, int]] = []
    for row in ws.iter_rows(min_row=1, max_row=40):
        values = [cell.value for cell in row]
        player_col, team_col = find_role_columns(values)
        if player_col is None:
            continue
        nonblank = sum(1 for value in values if text_value(value))
        score = nonblank + (100 if team_col is not None else 20)
        if audit_header_row and row[0].row == audit_header_row:
            score += 15
        candidates.append((score, row[0].row, team_col is not None, nonblank))

    if not candidates:
        if audit_header_row:
            add_issue(
                issues,
                "error",
                "unclear_header_row",
                "No Player/Team-style header row could be detected; audit header was not used without a player column.",
                sheetName=sheet_name,
                auditHeaderRow=audit_header_row,
            )
        return None

    candidates.sort(reverse=True)
    best_score, best_row, has_team, _ = candidates[0]
    tied = [candidate for candidate in candidates if candidate[0] == best_score]
    if len(tied) > 1:
        add_issue(
            issues,
            "warning",
            "ambiguous_header_row",
            "Multiple possible header rows had the same detection score; the first best row was used.",
            sheetName=sheet_name,
            chosenHeaderRow=best_row,
            candidateRows=[candidate[1] for candidate in tied],
        )
    if not has_team:
        add_issue(
            issues,
            "warning",
            "missing_team_column",
            "A Player header was detected but no Team header was found; team values will be preserved as blank.",
            sheetName=sheet_name,
            headerRow=best_row,
        )
    return best_row


def grouped_labels_for_headers(ws: Any, header_row: int, width: int) -> list[str]:
    if header_row <= 1:
        return [""] * width
    group_values = row_values(ws, header_row - 1)
    labels: list[str] = []
    current = ""
    for idx in range(width):
        value = group_values[idx] if idx < len(group_values) else None
        text = text_value(value).replace("\n", " ").strip()
        if text:
            current = text
        labels.append(current)
    return labels


def build_clean_header_overrides(
    *,
    sheet_name: str,
    headers: list[Any],
    group_labels: list[str],
    issues: list[dict[str, Any]],
) -> tuple[dict[int, str], dict[int, list[str]]]:
    overrides: dict[int, str] = {}
    notes: dict[int, list[str]] = defaultdict(list)
    reviewed_replacements = REVIEWED_HEADER_REPLACEMENTS.get(sheet_name, {})

    for idx, replacement in reviewed_replacements.items():
        overrides[idx] = snake_case(replacement)
        notes[idx].append(f"inferred_header:{replacement}")

    for idx, original in enumerate(headers, start=1):
        original_text = text_value(original).replace("\n", " ").strip()
        previous_text = text_value(headers[idx - 2]).replace("\n", " ").strip() if idx > 1 else ""
        next_text = text_value(headers[idx]).replace("\n", " ").strip() if idx < len(headers) else ""
        group_label = group_labels[idx - 1] if idx - 1 < len(group_labels) else ""

        if idx == 1 and not original_text and any(snake_case(header) in PLAYER_HEADER_NAMES for header in headers[1:4]):
            overrides[idx] = "rank"
            notes[idx].append("blank_leading_rank_header_inferred")

        if original_text == "#":
            overrides[idx] = "rank"
            notes[idx].append("rank_header_normalized")

        if is_time_formatted_three_header(original):
            if "3 Point Field Goals" in group_label or previous_text.upper() == "3FG FREQ":
                inferred = "three_fgm"
                readable = "3FGM"
            elif next_text.upper() == "3PA":
                inferred = "three_pm"
                readable = "3PM"
            else:
                inferred = "three_pm"
                readable = "3PM"
            overrides[idx] = inferred
            notes[idx].append(f"time_formatted_header_inferred_as:{readable}")
            add_issue(
                issues,
                "info",
                "time_formatted_header_normalized",
                "A time-formatted header was preserved raw and mapped to a usable cleaned stat name.",
                sheetName=sheet_name,
                columnIndex=idx,
                originalColumnName=original_text,
                cleanedColumnName=inferred,
                groupLabel=group_label,
            )

        if sheet_name in GROUPED_HEADER_SHEETS and group_label and original_text:
            role_name = snake_case(original_text)
            if role_name not in PLAYER_HEADER_NAMES | TEAM_HEADER_NAMES and role_name != "age":
                grouped_name = f"{snake_case(group_label)}_{role_name}"
                if idx not in overrides:
                    overrides[idx] = grouped_name
                    notes[idx].append(f"grouped_header_prefix:{group_label}")

    return overrides, notes


def prepare_columns(
    *,
    sheet_name: str,
    headers: list[Any],
    player_col: int,
    team_col: int | None,
    cleaned_overrides: dict[int, str] | None = None,
    header_notes: dict[int, list[str]] | None = None,
    issues: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    column_entries: list[dict[str, Any]] = []
    stat_columns: list[dict[str, Any]] = []
    seen_clean_names: Counter[str] = Counter()
    seen_original_stat_names: Counter[str] = Counter()
    for idx, original in enumerate(headers, start=1):
        original_name = text_value(original)
        excel_column = excel_column_letter(idx)
        role = "stat"
        imported = True
        notes: list[str] = list((header_notes or {}).get(idx, []))
        cleaned = (cleaned_overrides or {}).get(idx) or snake_case(original_name)
        if idx == player_col:
            role = "player"
        elif idx == team_col:
            role = "team"
        elif not original_name and idx not in (cleaned_overrides or {}):
            role = "skipped_blank_header"
            imported = False
            notes.append("blank_header_not_imported")
            add_issue(
                issues,
                "warning",
                "blank_header_column_skipped",
                "A blank header column was skipped during ingestion.",
                sheetName=sheet_name,
                columnIndex=idx,
                excelColumn=excel_column,
            )

        if imported and role == "stat":
            if original_name:
                seen_original_stat_names[original_name] += 1
                if seen_original_stat_names[original_name] > 1:
                    add_issue(
                        issues,
                        "warning",
                        "duplicate_original_column_name",
                        "A duplicate original column name exists in this sheet; cleaned names remain distinct where possible.",
                        sheetName=sheet_name,
                        originalColumnName=original_name,
                        duplicateIndex=seen_original_stat_names[original_name],
                        excelColumn=excel_column,
                    )
            seen_clean_names[cleaned] += 1
            if seen_clean_names[cleaned] > 1:
                deduped = f"{cleaned}_{seen_clean_names[cleaned]}"
                add_issue(
                    issues,
                    "warning",
                    "duplicate_clean_column_name",
                    "A duplicate cleaned column name was suffixed instead of overwritten.",
                    sheetName=sheet_name,
                    originalColumnName=original_name,
                    cleanedColumnName=cleaned,
                    dedupedColumnName=deduped,
                    excelColumn=excel_column,
                )
                cleaned = deduped
                notes.append("cleaned_column_name_deduped")

        entry = {
            "source_sheet": sheet_name,
            "column_index": idx,
            "excel_column": excel_column,
            "original_column_name": original_name,
            "cleaned_column_name": cleaned,
            "role": role,
            "imported": imported,
            "notes": notes,
        }
        column_entries.append(entry)
        if imported and role == "stat":
            stat_columns.append(entry)
    return column_entries, stat_columns


def create_database(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    for sidecar_suffix in ("-wal", "-shm"):
        sidecar = path.with_name(f"{path.name}{sidecar_suffix}")
        if sidecar.exists():
            sidecar.unlink()
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA journal_mode = DELETE")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.executescript(
        """
        CREATE TABLE import_metadata (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );

        CREATE TABLE imported_sheets (
          source_sheet TEXT PRIMARY KEY,
          stat_category TEXT NOT NULL,
          header_row INTEGER NOT NULL,
          source_row_count INTEGER NOT NULL,
          source_column_count INTEGER NOT NULL,
          imported_data_rows INTEGER NOT NULL,
          stat_rows_created INTEGER NOT NULL
        );

        CREATE TABLE column_dictionary (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          source_sheet TEXT NOT NULL,
          stat_category TEXT NOT NULL,
          column_index INTEGER NOT NULL,
          excel_column TEXT NOT NULL,
          original_column_name TEXT NOT NULL,
          cleaned_column_name TEXT NOT NULL,
          role TEXT NOT NULL,
          imported INTEGER NOT NULL,
          notes TEXT
        );

        CREATE TABLE players (
          player_slug TEXT PRIMARY KEY,
          player_name TEXT NOT NULL,
          season TEXT NOT NULL,
          season_type TEXT NOT NULL,
          primary_team TEXT,
          teams_json TEXT NOT NULL,
          name_variants_json TEXT NOT NULL,
          source_sheets_json TEXT NOT NULL,
          stat_rows INTEGER NOT NULL,
          profile_path TEXT NOT NULL
        );

        CREATE TABLE stat_values (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          raw_player_name TEXT,
          player_name TEXT NOT NULL,
          player_slug TEXT NOT NULL,
          team TEXT,
          season TEXT NOT NULL,
          season_type TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          stat_category TEXT NOT NULL,
          original_column_name TEXT NOT NULL,
          cleaned_column_name TEXT NOT NULL,
          raw_value TEXT,
          raw_value_json TEXT,
          numeric_value REAL,
          import_notes TEXT,
          source_row_number INTEGER NOT NULL,
          source_column_letter TEXT NOT NULL
        );

        CREATE INDEX idx_stat_values_player ON stat_values(player_slug);
        CREATE INDEX idx_stat_values_sheet ON stat_values(source_sheet);
        CREATE INDEX idx_stat_values_column ON stat_values(cleaned_column_name);
        """
    )
    return conn


def write_json(path: Path, data: Any, *, compact: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if compact:
        path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    else:
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def runtime_value(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "raw": record["raw_value"],
        "numeric": record["numeric_value"],
        "original_column_name": record["original_column_name"],
    }


def build_runtime_summary(profile: dict[str, Any], primary_team: str | None) -> dict[str, Any]:
    sheets: dict[str, dict[str, dict[str, Any]]] = {}
    for record in profile["stats"]:
        sheet_name = record["source_sheet"]
        if sheet_name not in RUNTIME_SUMMARY_SHEETS:
            continue
        sheet = sheets.setdefault(sheet_name, {})
        cleaned_name = record["cleaned_column_name"]
        if cleaned_name not in sheet:
            sheet[cleaned_name] = runtime_value(record)

    return {
        "player_name": profile["player_name"],
        "player_slug": profile["player_slug"],
        "season": SEASON,
        "season_type": SEASON_TYPE,
        "primary_team": primary_team,
        "teams": sorted(team for team in profile["teams"] if team),
        "source_sheets": sorted(profile["source_sheets"]),
        "sheets": sheets,
    }


def database_url_from_env() -> str | None:
    value = os.environ.get("DATABASE_URL", "").strip()
    return value or None


def normalized_lookup_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", text_value(value)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def int_or_none(value: Any) -> int | None:
    if value is None:
        return None
    try:
        if isinstance(value, str):
            value = value.strip()
            if value.lower() in MISSING_MARKERS:
                return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in MISSING_MARKERS:
            return None
        value = value.replace(",", "")
        if value.endswith("%"):
            value = value[:-1]
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def percent_or_none(value: Any) -> float | None:
    parsed = float_or_none(value)
    if parsed is None:
        return None
    return parsed / 100 if abs(parsed) > 1 else parsed


def table_rows(snapshot: dict[str, Any], table_name: str) -> list[dict[str, Any]]:
    table = snapshot.get("tables", {}).get(table_name) or {}
    headers = table.get("headers") or []
    rows = table.get("rows") or []
    return [dict(zip(headers, row)) for row in rows]


def load_official_reference(snapshot_path: Path) -> dict[str, Any]:
    if not snapshot_path.exists():
        return {
            "teams_by_id": {},
            "teams_by_abbr": {},
            "players_by_name": {},
            "players_by_name_team": {},
        }

    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    player_index_rows = table_rows(snapshot, "playerIndex")
    player_bio_rows = table_rows(snapshot, "playerBioStatsRegular")
    team_stat_rows = table_rows(snapshot, "teamStatsRegular")
    bio_by_id = {str(row.get("PLAYER_ID")): row for row in player_bio_rows if row.get("PLAYER_ID") is not None}

    teams_by_id: dict[str, dict[str, Any]] = {}
    for row in player_index_rows:
        team_id = str(row.get("TEAM_ID") or "").strip()
        abbr = text_value(row.get("TEAM_ABBREVIATION")).strip()
        if not team_id or team_id == "0" or not abbr:
            continue
        teams_by_id.setdefault(
            team_id,
            {
                "id": team_id,
                "slug": slugify(text_value(row.get("TEAM_SLUG")) or abbr),
                "abbreviation": abbr,
                "city": text_value(row.get("TEAM_CITY")) or abbr,
                "name": text_value(row.get("TEAM_NAME")) or abbr,
                "conference": None,
                "division": None,
                "primary_color": None,
                "secondary_color": None,
                "source": "official_snapshot",
            },
        )

    for row in team_stat_rows:
        team_id = str(row.get("TEAM_ID") or "").strip()
        full_name = text_value(row.get("TEAM_NAME")).strip()
        if not team_id or team_id in teams_by_id or not full_name:
            continue
        teams_by_id[team_id] = {
            "id": team_id,
            "slug": slugify(full_name),
            "abbreviation": slugify(full_name).upper()[:3],
            "city": full_name,
            "name": full_name,
            "conference": None,
            "division": None,
            "primary_color": None,
            "secondary_color": None,
            "source": "official_snapshot",
        }

    teams_by_abbr = {team["abbreviation"]: team for team in teams_by_id.values() if team.get("abbreviation")}
    players_by_name: dict[str, dict[str, Any]] = {}
    players_by_name_team: dict[tuple[str, str], dict[str, Any]] = {}
    for row in player_index_rows:
        person_id = str(row.get("PERSON_ID") or "").strip()
        first = text_value(row.get("PLAYER_FIRST_NAME")).strip()
        last = text_value(row.get("PLAYER_LAST_NAME")).strip()
        player_name = f"{first} {last}".strip()
        if not player_name:
            continue
        team_abbr = text_value(row.get("TEAM_ABBREVIATION")).strip()
        bio = bio_by_id.get(person_id, {})
        reference_row = {
            "nba_player_id": person_id or None,
            "app_player_id": person_id or None,
            "player_name": player_name,
            "normalized_player_name": normalized_lookup_key(player_name),
            "team_abbreviation": team_abbr or None,
            "team_id": str(row.get("TEAM_ID") or "").strip() or None,
            "position": text_value(row.get("POSITION")).strip() or None,
            "height": text_value(row.get("HEIGHT") or bio.get("PLAYER_HEIGHT")).strip() or None,
            "height_inches": float_or_none(bio.get("PLAYER_HEIGHT_INCHES")),
            "weight": int_or_none(row.get("WEIGHT") or bio.get("PLAYER_WEIGHT")),
            "age": int_or_none(bio.get("AGE")),
            "college": text_value(row.get("COLLEGE") or bio.get("COLLEGE")).strip() or None,
            "country": text_value(row.get("COUNTRY") or bio.get("COUNTRY")).strip() or None,
            "jersey_number": text_value(row.get("JERSEY_NUMBER")).strip() or None,
            "headshot_url": f"https://cdn.nba.com/headshots/nba/latest/1040x760/{person_id}.png" if person_id else None,
        }
        lookup_name = normalized_lookup_key(player_name)
        players_by_name.setdefault(lookup_name, reference_row)
        if team_abbr:
            players_by_name_team[(lookup_name, team_abbr)] = reference_row

    return {
        "teams_by_id": teams_by_id,
        "teams_by_abbr": teams_by_abbr,
        "players_by_name": players_by_name,
        "players_by_name_team": players_by_name_team,
    }


def fallback_team_row(abbreviation: str) -> dict[str, Any]:
    safe_abbreviation = abbreviation.strip() or "UNK"
    if safe_abbreviation.upper() == "TOT":
        city = "Multiple"
        name = "Teams"
    else:
        city = safe_abbreviation
        name = safe_abbreviation
    return {
        "id": safe_abbreviation,
        "slug": slugify(safe_abbreviation),
        "abbreviation": safe_abbreviation,
        "city": city,
        "name": name,
        "conference": None,
        "division": None,
        "primary_color": None,
        "secondary_color": None,
        "source": "excel_fallback",
    }


def find_player_reference(player: dict[str, Any], official: dict[str, Any]) -> dict[str, Any] | None:
    name_key = normalized_lookup_key(player["player_name"])
    primary_team = player.get("primary_team")
    if primary_team:
        by_team = official["players_by_name_team"].get((name_key, primary_team))
        if by_team:
            return by_team
    return official["players_by_name"].get(name_key)


def source_sheet_value(summary: dict[str, Any], source_sheet: str, cleaned_column_name: str) -> dict[str, Any] | None:
    return summary.get("sheets", {}).get(source_sheet, {}).get(cleaned_column_name)


def summary_numeric(summary: dict[str, Any], source_sheet: str, cleaned_column_name: str) -> float | None:
    cell = source_sheet_value(summary, source_sheet, cleaned_column_name)
    if not cell:
        return None
    return float_or_none(cell.get("numeric"))


def summary_percent(summary: dict[str, Any], source_sheet: str, cleaned_column_name: str) -> float | None:
    cell = source_sheet_value(summary, source_sheet, cleaned_column_name)
    if not cell:
        return None
    return percent_or_none(cell.get("numeric"))


def stat_row_fingerprint(row: tuple[Any, ...]) -> str:
    payload = {
        "player_slug": row[2],
        "team": row[3],
        "season": row[4],
        "season_type": row[5],
        "source_sheet": row[6],
        "stat_category": row[7],
        "original_column_name": row[8],
        "cleaned_column_name": row[9],
        "raw_value_json": row[11],
        "source_row_number": row[14],
        "source_column_letter": row[15],
    }
    encoded = json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def batched(rows: list[Any], size: int) -> Any:
    for start in range(0, len(rows), size):
        yield rows[start : start + size]


def write_postgres_outputs(
    *,
    args: argparse.Namespace,
    generated_at: str,
    workbook_path: Path,
    audit: dict[str, Any],
    imported_sheets: list[str],
    skipped_sheet_names: list[str],
    failed_sheet_names: list[str],
    all_column_entries: list[dict[str, Any]],
    all_stat_rows: list[tuple[Any, ...]],
    players_json: list[dict[str, Any]],
    profile_outputs: dict[str, dict[str, Any]],
    runtime_summaries: list[dict[str, Any]],
    issue_log: dict[str, Any],
) -> dict[str, Any]:
    database_url = database_url_from_env()
    if not database_url:
        raise RuntimeError("Postgres write requested with --write-postgres, but DATABASE_URL is not set.")

    try:
        import psycopg
        from psycopg.types.json import Jsonb
    except ImportError as exc:
        raise RuntimeError(
            "Postgres write requested, but psycopg is not installed. Install Python dependencies from requirements.txt."
        ) from exc

    run_id = str(uuid.uuid4())
    official = load_official_reference(args.official_snapshot)
    runtime_by_slug = {summary["player_slug"]: summary for summary in runtime_summaries}

    teams_by_abbr: dict[str, dict[str, Any]] = dict(official["teams_by_abbr"])
    for player in players_json:
        for team_abbr in player.get("teams", []):
            if team_abbr and team_abbr not in teams_by_abbr:
                teams_by_abbr[team_abbr] = fallback_team_row(team_abbr)
    team_rows_by_id = {team["id"]: team for team in teams_by_abbr.values()}
    team_id_by_abbr = {team["abbreviation"]: team["id"] for team in team_rows_by_id.values()}

    player_rows = []
    profile_rows = []
    season_summary_rows = []
    for player in players_json:
        reference = find_player_reference(player, official) or {}
        primary_team = player.get("primary_team")
        primary_team_id = team_id_by_abbr.get(primary_team) if primary_team else None
        player_rows.append(
            {
                "player_slug": player["player_slug"],
                "nba_player_id": reference.get("nba_player_id"),
                "app_player_id": reference.get("app_player_id"),
                "player_name": player["player_name"],
                "normalized_player_name": normalized_lookup_key(player["player_name"]),
                "primary_team_id": primary_team_id,
                "primary_team_abbreviation": primary_team,
                "position": reference.get("position"),
                "height": reference.get("height"),
                "height_inches": reference.get("height_inches"),
                "weight": reference.get("weight"),
                "age": reference.get("age"),
                "college": reference.get("college"),
                "country": reference.get("country"),
                "jersey_number": reference.get("jersey_number"),
                "headshot_url": reference.get("headshot_url"),
            }
        )
        profile = profile_outputs[player["player_slug"]]
        profile_json = {key: value for key, value in profile.items() if key != "stats"}
        profile_json["profile_path"] = player["profile_path"]
        profile_rows.append(
            {
                "player_slug": player["player_slug"],
                "season": SEASON,
                "season_type": SEASON_TYPE,
                "primary_team": primary_team,
                "teams": player.get("teams", []),
                "name_variants": player.get("name_variants", []),
                "source_sheets": player.get("source_sheets", []),
                "stat_rows": player.get("stat_rows", 0),
                "profile_json": profile_json,
            }
        )

        summary = runtime_by_slug.get(player["player_slug"], {})
        games = int_or_none(summary_numeric(summary, "General - Traditional", "gp"))
        season_summary_rows.append(
            {
                "player_slug": player["player_slug"],
                "team_id": primary_team_id,
                "season": SEASON,
                "season_type": SEASON_TYPE,
                "position": reference.get("position"),
                "games": games,
                "minutes": summary_numeric(summary, "General - Traditional", "min"),
                "pts": summary_numeric(summary, "General - Traditional", "pts"),
                "reb": summary_numeric(summary, "General - Traditional", "reb"),
                "ast": summary_numeric(summary, "General - Traditional", "ast"),
                "stl": summary_numeric(summary, "General - Traditional", "stl"),
                "blk": summary_numeric(summary, "General - Traditional", "blk"),
                "tov": summary_numeric(summary, "General - Traditional", "tov"),
                "fg_pct": summary_percent(summary, "General - Traditional", "fg_pct"),
                "three_pct": summary_percent(summary, "General - Traditional", "three_p_pct"),
                "ft_pct": summary_percent(summary, "General - Traditional", "ft_pct"),
                "ts_pct": summary_percent(summary, "General - Advanced", "ts_pct"),
                "efg_pct": summary_percent(summary, "General - Advanced", "efg_pct"),
                "usage_rate": summary_percent(summary, "General - Advanced", "usg_pct"),
                "ast_pct": summary_percent(summary, "General - Advanced", "ast_pct"),
                "reb_pct": summary_percent(summary, "General - Advanced", "reb_pct"),
                "turnover_rate": summary_percent(summary, "General - Advanced", "to_ratio"),
                "off_rating": summary_numeric(summary, "General - Advanced", "offrtg"),
                "def_rating": summary_numeric(summary, "General - Advanced", "defrtg"),
                "net_rating": summary_numeric(summary, "General - Advanced", "netrtg"),
                "pie": summary_percent(summary, "General - Advanced", "pie"),
                "summary_json": summary,
            }
        )

    stat_categories = sorted({(entry["source_sheet"], entry["stat_category"]) for entry in all_column_entries})
    stat_value_rows = []
    for row in all_stat_rows:
        team_id = team_id_by_abbr.get(row[3]) if row[3] else None
        stat_value_rows.append(
            (
                run_id,
                row[0],
                row[1],
                row[2],
                row[3],
                team_id,
                row[4],
                row[5],
                row[6],
                row[7],
                row[8],
                row[9],
                row[10],
                Jsonb(json.loads(row[11])),
                row[12],
                row[13],
                row[14],
                row[15],
                stat_row_fingerprint(row),
            )
        )

    run_metadata = {
        "generated_at": generated_at,
        "json_outputs_preserved": True,
        "sheets_imported": imported_sheets,
        "sheets_skipped": skipped_sheet_names,
        "sheets_failed": failed_sheet_names,
        "source": "scripts/ingest_nba_excel.py",
    }
    batch_size = max(1, args.postgres_batch_size)

    try:
        with psycopg.connect(database_url) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO ingestion_runs (
                      id, source_workbook_path, source_workbook_sha256, season, season_type,
                      status, sheets_found, sheets_imported, sheets_skipped, sheets_failed,
                      unique_players, stat_rows_created, issues_logged, metadata
                    )
                    VALUES (%s, %s, %s, %s, %s, 'running', %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        run_id,
                        str(workbook_path),
                        sha256(workbook_path),
                        SEASON,
                        SEASON_TYPE,
                        len(audit["sheets"]),
                        len(imported_sheets),
                        len(skipped_sheet_names),
                        len(failed_sheet_names),
                        len(players_json),
                        len(all_stat_rows),
                        len(issue_log["issues"]),
                        Jsonb(run_metadata),
                    ),
                )
                cur.executemany(
                    """
                    INSERT INTO teams (
                      id, slug, abbreviation, city, name, conference, division,
                      primary_color, secondary_color, source, updated_at
                    )
                    VALUES (
                      %(id)s, %(slug)s, %(abbreviation)s, %(city)s, %(name)s, %(conference)s, %(division)s,
                      %(primary_color)s, %(secondary_color)s, %(source)s, now()
                    )
                    ON CONFLICT (id) DO UPDATE SET
                      slug = EXCLUDED.slug,
                      abbreviation = EXCLUDED.abbreviation,
                      city = EXCLUDED.city,
                      name = EXCLUDED.name,
                      conference = EXCLUDED.conference,
                      division = EXCLUDED.division,
                      primary_color = EXCLUDED.primary_color,
                      secondary_color = EXCLUDED.secondary_color,
                      source = EXCLUDED.source,
                      updated_at = now()
                    """,
                    list(team_rows_by_id.values()),
                )
                cur.executemany(
                    """
                    INSERT INTO players (
                      player_slug, nba_player_id, app_player_id, player_name, normalized_player_name,
                      primary_team_id, primary_team_abbreviation, position, height, height_inches, weight,
                      age, college, country, jersey_number, headshot_url, active, updated_at
                    )
                    VALUES (
                      %(player_slug)s, %(nba_player_id)s, %(app_player_id)s, %(player_name)s, %(normalized_player_name)s,
                      %(primary_team_id)s, %(primary_team_abbreviation)s, %(position)s, %(height)s, %(height_inches)s,
                      %(weight)s, %(age)s, %(college)s, %(country)s, %(jersey_number)s, %(headshot_url)s, true, now()
                    )
                    ON CONFLICT (player_slug) DO UPDATE SET
                      nba_player_id = EXCLUDED.nba_player_id,
                      app_player_id = EXCLUDED.app_player_id,
                      player_name = EXCLUDED.player_name,
                      normalized_player_name = EXCLUDED.normalized_player_name,
                      primary_team_id = EXCLUDED.primary_team_id,
                      primary_team_abbreviation = EXCLUDED.primary_team_abbreviation,
                      position = EXCLUDED.position,
                      height = EXCLUDED.height,
                      height_inches = EXCLUDED.height_inches,
                      weight = EXCLUDED.weight,
                      age = EXCLUDED.age,
                      college = EXCLUDED.college,
                      country = EXCLUDED.country,
                      jersey_number = EXCLUDED.jersey_number,
                      headshot_url = EXCLUDED.headshot_url,
                      active = true,
                      updated_at = now()
                    """,
                    player_rows,
                )
                cur.executemany(
                    """
                    INSERT INTO stat_categories (source_sheet, stat_category)
                    VALUES (%s, %s)
                    ON CONFLICT (source_sheet, stat_category) DO NOTHING
                    """,
                    stat_categories,
                )
                cur.executemany(
                    """
                    INSERT INTO column_dictionary (
                      ingestion_run_id, source_sheet, stat_category, column_index, excel_column,
                      original_column_name, cleaned_column_name, role, imported, notes
                    )
                    VALUES (
                      %(ingestion_run_id)s, %(source_sheet)s, %(stat_category)s, %(column_index)s, %(excel_column)s,
                      %(original_column_name)s, %(cleaned_column_name)s, %(role)s, %(imported)s, %(notes)s
                    )
                    ON CONFLICT (ingestion_run_id, source_sheet, column_index, cleaned_column_name) DO NOTHING
                    """,
                    [
                        {
                            "ingestion_run_id": run_id,
                            **entry,
                            "notes": "; ".join(entry["notes"]) if entry["notes"] else None,
                        }
                        for entry in all_column_entries
                    ],
                )
                cur.executemany(
                    """
                    INSERT INTO player_profiles (
                      ingestion_run_id, player_slug, season, season_type, primary_team, teams,
                      name_variants, source_sheets, stat_rows, profile_json
                    )
                    VALUES (
                      %(ingestion_run_id)s, %(player_slug)s, %(season)s, %(season_type)s, %(primary_team)s, %(teams)s,
                      %(name_variants)s, %(source_sheets)s, %(stat_rows)s, %(profile_json)s
                    )
                    ON CONFLICT (player_slug, season, season_type, ingestion_run_id) DO UPDATE SET
                      primary_team = EXCLUDED.primary_team,
                      teams = EXCLUDED.teams,
                      name_variants = EXCLUDED.name_variants,
                      source_sheets = EXCLUDED.source_sheets,
                      stat_rows = EXCLUDED.stat_rows,
                      profile_json = EXCLUDED.profile_json
                    """,
                    [
                        {
                            "ingestion_run_id": run_id,
                            **row,
                            "teams": Jsonb(row["teams"]),
                            "name_variants": Jsonb(row["name_variants"]),
                            "source_sheets": Jsonb(row["source_sheets"]),
                            "profile_json": Jsonb(row["profile_json"]),
                        }
                        for row in profile_rows
                    ],
                    returning=False,
                )
                cur.executemany(
                    """
                    INSERT INTO player_season_summaries (
                      ingestion_run_id, player_slug, team_id, season, season_type, position, games,
                      minutes, pts, reb, ast, stl, blk, tov, fg_pct, three_pct, ft_pct, ts_pct,
                      efg_pct, usage_rate, ast_pct, reb_pct, turnover_rate, off_rating, def_rating,
                      net_rating, pie, summary_json
                    )
                    VALUES (
                      %(ingestion_run_id)s, %(player_slug)s, %(team_id)s, %(season)s, %(season_type)s, %(position)s, %(games)s,
                      %(minutes)s, %(pts)s, %(reb)s, %(ast)s, %(stl)s, %(blk)s, %(tov)s, %(fg_pct)s, %(three_pct)s,
                      %(ft_pct)s, %(ts_pct)s, %(efg_pct)s, %(usage_rate)s, %(ast_pct)s, %(reb_pct)s,
                      %(turnover_rate)s, %(off_rating)s, %(def_rating)s, %(net_rating)s, %(pie)s, %(summary_json)s
                    )
                    ON CONFLICT (player_slug, season, season_type, ingestion_run_id) DO UPDATE SET
                      team_id = EXCLUDED.team_id,
                      position = EXCLUDED.position,
                      games = EXCLUDED.games,
                      minutes = EXCLUDED.minutes,
                      pts = EXCLUDED.pts,
                      reb = EXCLUDED.reb,
                      ast = EXCLUDED.ast,
                      stl = EXCLUDED.stl,
                      blk = EXCLUDED.blk,
                      tov = EXCLUDED.tov,
                      fg_pct = EXCLUDED.fg_pct,
                      three_pct = EXCLUDED.three_pct,
                      ft_pct = EXCLUDED.ft_pct,
                      ts_pct = EXCLUDED.ts_pct,
                      efg_pct = EXCLUDED.efg_pct,
                      usage_rate = EXCLUDED.usage_rate,
                      ast_pct = EXCLUDED.ast_pct,
                      reb_pct = EXCLUDED.reb_pct,
                      turnover_rate = EXCLUDED.turnover_rate,
                      off_rating = EXCLUDED.off_rating,
                      def_rating = EXCLUDED.def_rating,
                      net_rating = EXCLUDED.net_rating,
                      pie = EXCLUDED.pie,
                      summary_json = EXCLUDED.summary_json
                    """,
                    [{**row, "ingestion_run_id": run_id, "summary_json": Jsonb(row["summary_json"])} for row in season_summary_rows],
                    returning=False,
                )
                # Make the run metadata and small lookup tables durable before the
                # large fact load. Current views only expose succeeded runs, so
                # these rows remain invisible until finalization.
                conn.commit()
                for batch in batched(stat_value_rows, batch_size):
                    with cur.copy(
                        """
                        COPY player_stat_values (
                          ingestion_run_id, raw_player_name, player_name, player_slug, team, team_id,
                          season, season_type, source_sheet, stat_category, original_column_name,
                          cleaned_column_name, raw_value, raw_value_json, numeric_value, import_notes,
                          source_row_number, source_column_letter, row_fingerprint
                        )
                        FROM STDIN
                        """
                    ) as copy:
                        for stat_row in batch:
                            copy.write_row(stat_row)
                    conn.commit()
                cur.executemany(
                    """
                    INSERT INTO data_issues (ingestion_run_id, severity, type, message, details)
                    VALUES (%(ingestion_run_id)s, %(severity)s, %(type)s, %(message)s, %(details)s)
                    """,
                    [
                        {**issue, "ingestion_run_id": run_id, "details": Jsonb(issue.get("details", {}))}
                        for issue in issue_log["issues"]
                    ],
                    returning=False,
                )
                cur.execute(
                    """
                    UPDATE ingestion_runs
                    SET status = 'succeeded', finished_at = now()
                    WHERE id = %s
                    """,
                    (run_id,),
                )
    except Exception as exc:
        try:
            with psycopg.connect(database_url) as conn:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM ingestion_runs WHERE id = %s", (run_id,))
                    cur.execute(
                        """
                        INSERT INTO ingestion_runs (
                          id, source_workbook_path, source_workbook_sha256, season, season_type,
                          status, started_at, finished_at, metadata
                        )
                        VALUES (%s, %s, %s, %s, %s, 'failed', now(), now(), %s)
                        ON CONFLICT (id) DO UPDATE SET
                          status = 'failed',
                          finished_at = now(),
                          metadata = EXCLUDED.metadata
                        """,
                        (
                            run_id,
                            str(workbook_path),
                            sha256(workbook_path),
                            SEASON,
                            SEASON_TYPE,
                            Jsonb({"error": str(exc), **run_metadata}),
                        ),
                    )
        except Exception:
            pass
        raise RuntimeError(f"Postgres ingestion failed and was rolled back: {exc}") from exc

    return {
        "requested": True,
        "written": True,
        "ingestion_run_id": run_id,
        "tables_written": [
            "ingestion_runs",
            "teams",
            "players",
            "player_profiles",
            "player_season_summaries",
            "player_stat_values",
            "stat_categories",
            "column_dictionary",
            "data_issues",
        ],
        "teams_written": len(team_rows_by_id),
        "players_written": len(player_rows),
        "stat_rows_written": len(stat_value_rows),
        "issues_written": len(issue_log["issues"]),
    }


def import_workbook(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    from audit_nba_excel import build_audit
    from openpyxl import load_workbook

    generated_at = now_iso()
    audit = build_audit(workbook_path)
    sheet_decisions = load_sheet_decisions(audit)
    importable_sheet_names = [
        sheet["sheetName"]
        for sheet in audit["sheets"]
        if sheet["decision"] in {"import", "review"} or sheet["sheetName"] in REVIEWED_IMPORT_SHEETS
    ]
    skipped_sheet_names = [
        sheet["sheetName"]
        for sheet in audit["sheets"]
        if sheet["decision"] == "skip" and sheet["sheetName"] not in REVIEWED_IMPORT_SHEETS
    ]
    issues: list[dict[str, Any]] = []
    for sheet_name in importable_sheet_names:
        review_note = REVIEWED_IMPORT_SHEETS.get(sheet_name)
        if sheet_decisions[sheet_name]["decision"] == "review":
            add_issue(
                issues,
                "info",
                "reviewed_sheet_imported",
                "A sheet marked review by the audit was explicitly imported after deterministic review.",
                sheetName=sheet_name,
                reviewNote=review_note or "Imported because the sheet has a detectable NBA.com player table header.",
                auditIssues=[item["type"] for item in sheet_decisions[sheet_name]["issues"]],
            )
    for sheet_name in skipped_sheet_names:
        sheet = sheet_decisions[sheet_name]
        add_issue(
            issues,
            "warning",
            "sheet_skipped_not_importable",
            "Sheet was skipped because the audit did not mark it as a clean import candidate.",
            sheetName=sheet_name,
            decision=sheet["decision"],
            auditIssues=[item["type"] for item in sheet["issues"]],
        )

    conn = create_database(args.sqlite)
    metadata_rows = [
        ("generated_at", generated_at),
        ("source_workbook", str(workbook_path)),
        ("source_workbook_sha256", sha256(workbook_path)),
        ("season", SEASON),
        ("season_type", SEASON_TYPE),
        ("audit_summary", json.dumps(audit["summary"], ensure_ascii=False)),
    ]
    conn.executemany("INSERT INTO import_metadata (key, value) VALUES (?, ?)", metadata_rows)

    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    all_column_entries: list[dict[str, Any]] = []
    all_stat_rows: list[tuple[Any, ...]] = []
    players: dict[str, dict[str, Any]] = {}
    player_profiles: dict[str, dict[str, Any]] = {}
    profile_outputs: dict[str, dict[str, Any]] = {}
    imported_sheets: list[str] = []
    failed_sheet_names: list[str] = []
    sheet_summaries: list[dict[str, Any]] = []
    missing_value_counts: Counter[tuple[str, str, str]] = Counter()
    nonnumeric_counts: Counter[tuple[str, str, str]] = Counter()
    normalized_value_counts: Counter[tuple[str, str, str, str]] = Counter()
    player_name_reformat_counts: Counter[str] = Counter()
    player_name_variant_counts: Counter[tuple[str, str]] = Counter()
    missing_team_counts: Counter[str] = Counter()
    missing_team_samples: dict[str, list[dict[str, Any]]] = defaultdict(list)
    duplicate_entity_counts: Counter[tuple[str, str, str]] = Counter()
    duplicate_entity_samples: dict[tuple[str, str, str], list[int]] = defaultdict(list)

    try:
        for sheet_name in importable_sheet_names:
            try:
                sheet_meta = sheet_decisions[sheet_name]
                ws = wb[sheet_name]
                header_row = infer_header_row(ws, sheet_name, sheet_meta["likelyHeaderRow"], issues)
                if not header_row:
                    failed_sheet_names.append(sheet_name)
                    add_issue(
                        issues,
                        "error",
                        "failed_sheet_unclear_header",
                        "Sheet could not be imported because no clear player table header was detected.",
                        sheetName=sheet_name,
                    )
                    continue

                headers = row_values(ws, header_row)
                group_labels = grouped_labels_for_headers(ws, header_row, len(headers)) if sheet_name in GROUPED_HEADER_SHEETS else [""] * len(headers)
                cleaned_overrides, header_notes = build_clean_header_overrides(
                    sheet_name=sheet_name,
                    headers=headers,
                    group_labels=group_labels,
                    issues=issues,
                )
                player_col, team_col = find_role_columns(headers, cleaned_overrides)
                if player_col is None:
                    failed_sheet_names.append(sheet_name)
                    add_issue(
                        issues,
                        "error",
                        "failed_sheet_missing_player_column",
                        "Sheet could not be imported because no Player or VS PLAYER column was detected.",
                        sheetName=sheet_name,
                        headerRow=header_row,
                    )
                    continue

                stat_category = sheet_meta["possibleStatCategory"]
                column_entries, stat_columns = prepare_columns(
                    sheet_name=sheet_name,
                    headers=headers,
                    player_col=player_col,
                    team_col=team_col,
                    cleaned_overrides=cleaned_overrides,
                    header_notes=header_notes,
                    issues=issues,
                )
                for entry in column_entries:
                    entry["stat_category"] = stat_category
                all_column_entries.extend(column_entries)

                imported_data_rows = 0
                sheet_stat_rows = 0
                for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
                    values = [cell.value for cell in row]
                    if not any(text_value(value) for value in values):
                        continue
                    source_row_number = row[0].row
                    player_raw = values[player_col - 1] if player_col - 1 < len(values) else None
                    team_raw = values[team_col - 1] if team_col and team_col - 1 < len(values) else None
                    raw_player_name = text_value(player_raw)
                    player_name, player_name_note = canonical_player_name(player_raw)
                    team = text_value(team_raw)
                    if not player_name:
                        add_issue(
                            issues,
                            "error",
                            "row_skipped_missing_player",
                            "A data row was skipped because player name is blank.",
                            sheetName=sheet_name,
                            rowNumber=source_row_number,
                        )
                        continue
                    if player_name_note:
                        player_name_reformat_counts[sheet_name] += 1
                    if not team:
                        missing_team_counts[sheet_name] += 1
                        if len(missing_team_samples[sheet_name]) < 10:
                            missing_team_samples[sheet_name].append({"rowNumber": source_row_number, "playerName": player_name})

                    imported_data_rows += 1
                    player_slug = slugify(player_name)
                    entity_key = (sheet_name, player_slug, team)
                    duplicate_entity_counts[entity_key] += 1
                    if len(duplicate_entity_samples[entity_key]) < 5:
                        duplicate_entity_samples[entity_key].append(source_row_number)
                    if player_slug not in players:
                        players[player_slug] = {
                            "player_name": player_name,
                            "player_slug": player_slug,
                            "season": SEASON,
                            "season_type": SEASON_TYPE,
                            "teams": Counter(),
                            "source_sheets": set(),
                            "name_variants": set(),
                            "stat_rows": 0,
                        }
                    player = players[player_slug]
                    display_player_name = player["player_name"]
                    player["name_variants"].update({raw_player_name, player_name, display_player_name})
                    if player_name != display_player_name:
                        player_name_variant_counts[(display_player_name, player_name)] += 1
                    player["teams"][team] += 1
                    player["source_sheets"].add(sheet_name)
                    profile = player_profiles.setdefault(
                        player_slug,
                        {
                            "player_name": display_player_name,
                            "player_slug": player_slug,
                            "season": SEASON,
                            "season_type": SEASON_TYPE,
                            "teams": set(),
                            "source_sheets": set(),
                            "name_variants": set(),
                            "stats": [],
                        },
                    )
                    profile["teams"].add(team)
                    profile["source_sheets"].add(sheet_name)
                    profile["name_variants"].update({raw_player_name, player_name, display_player_name})

                    for column in stat_columns:
                        column_index = column["column_index"]
                        raw = values[column_index - 1] if column_index - 1 < len(values) else None
                        numeric_value, note = clean_numeric(
                            raw,
                            source_sheet=sheet_name,
                            cleaned_column_name=column["cleaned_column_name"],
                        )
                        notes = []
                        if note:
                            notes.append(note)
                            if numeric_value is None:
                                key = (sheet_name, column["cleaned_column_name"], column["original_column_name"])
                                if note == "blank_value":
                                    missing_value_counts[key] += 1
                                else:
                                    nonnumeric_counts[key] += 1
                            else:
                                normalized_value_counts[
                                    (
                                        sheet_name,
                                        column["cleaned_column_name"],
                                        column["original_column_name"],
                                        note,
                                    )
                                ] += 1
                        raw_json_value = json_safe(raw)
                        stat_record = {
                            "player_name": display_player_name,
                            "raw_player_name": raw_player_name,
                            "team": team,
                            "season": SEASON,
                            "season_type": SEASON_TYPE,
                            "source_sheet": sheet_name,
                            "stat_category": stat_category,
                            "original_column_name": column["original_column_name"],
                            "cleaned_column_name": column["cleaned_column_name"],
                            "raw_value": raw_json_value,
                            "numeric_value": numeric_value,
                            "import_notes": "; ".join(notes) if notes else None,
                        }
                        if PROFILE_STAT_LIMIT is None or len(profile["stats"]) < PROFILE_STAT_LIMIT:
                            profile["stats"].append(stat_record)
                        all_stat_rows.append(
                            (
                                raw_player_name,
                                display_player_name,
                                player_slug,
                                team,
                                SEASON,
                                SEASON_TYPE,
                                sheet_name,
                                stat_category,
                                column["original_column_name"],
                                column["cleaned_column_name"],
                                text_value(raw),
                                json.dumps(raw_json_value, ensure_ascii=False),
                                numeric_value,
                                stat_record["import_notes"],
                                source_row_number,
                                column["excel_column"],
                            )
                        )
                        player["stat_rows"] += 1
                        sheet_stat_rows += 1

                imported_sheets.append(sheet_name)
                sheet_summaries.append(
                    {
                        "source_sheet": sheet_name,
                        "stat_category": stat_category,
                        "header_row": header_row,
                        "source_row_count": sheet_meta["rowCount"],
                        "source_column_count": sheet_meta["columnCount"],
                        "imported_data_rows": imported_data_rows,
                        "stat_rows_created": sheet_stat_rows,
                    }
                )
            except Exception as exc:
                failed_sheet_names.append(sheet_name)
                add_issue(
                    issues,
                    "error",
                    "failed_sheet_import",
                    "Sheet import failed and was skipped after logging the exception.",
                    sheetName=sheet_name,
                    exceptionType=type(exc).__name__,
                    error=str(exc),
                )
    finally:
        wb.close()

    for (sheet_name, cleaned_name, original_name), count in missing_value_counts.items():
        add_issue(
            issues,
            "info",
            "blank_stat_values",
            "Blank stat values were preserved as null numeric values.",
            sheetName=sheet_name,
            cleanedColumnName=cleaned_name,
            originalColumnName=original_name,
            count=count,
        )
    for (sheet_name, cleaned_name, original_name), count in nonnumeric_counts.items():
        add_issue(
            issues,
            "warning",
            "non_numeric_stat_values",
            "Non-numeric stat values were preserved raw with null numeric values.",
            sheetName=sheet_name,
            cleanedColumnName=cleaned_name,
            originalColumnName=original_name,
            count=count,
        )
    for (sheet_name, cleaned_name, original_name, note), count in normalized_value_counts.items():
        add_issue(
            issues,
            "info",
            "normalized_numeric_values",
            "Raw values were preserved and deterministic numeric values were created for a typed stat field.",
            sheetName=sheet_name,
            cleanedColumnName=cleaned_name,
            originalColumnName=original_name,
            normalization=note,
            count=count,
        )
    for sheet_name, count in player_name_reformat_counts.items():
        add_issue(
            issues,
            "info",
            "player_name_reformatted",
            "Player names in Last, First format were deterministically normalized to First Last.",
            sheetName=sheet_name,
            count=count,
        )
    for (display_name, variant_name), count in player_name_variant_counts.items():
        add_issue(
            issues,
            "info",
            "player_name_variant_merged",
            "Player name variants were merged by normalized slug.",
            playerName=display_name,
            variantName=variant_name,
            count=count,
        )
    for sheet_name, count in missing_team_counts.items():
        add_issue(
            issues,
            "warning",
            "missing_team_values",
            "Rows with missing team values were imported with blank team fields; no team was guessed.",
            sheetName=sheet_name,
            count=count,
            sampleRows=missing_team_samples[sheet_name],
        )
    for (sheet_name, player_slug, team), count in duplicate_entity_counts.items():
        if count <= 1:
            continue
        add_issue(
            issues,
            "warning",
            "duplicate_player_team_rows",
            "The same player/team key appears more than once in a source sheet; all rows were preserved.",
            sheetName=sheet_name,
            playerSlug=player_slug,
            team=team,
            count=count,
            sampleRowNumbers=duplicate_entity_samples[(sheet_name, player_slug, team)],
        )

    conn.executemany(
        """
        INSERT INTO imported_sheets (
          source_sheet, stat_category, header_row, source_row_count, source_column_count,
          imported_data_rows, stat_rows_created
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                sheet["source_sheet"],
                sheet["stat_category"],
                sheet["header_row"],
                sheet["source_row_count"],
                sheet["source_column_count"],
                sheet["imported_data_rows"],
                sheet["stat_rows_created"],
            )
            for sheet in sheet_summaries
        ],
    )
    conn.executemany(
        """
        INSERT INTO column_dictionary (
          source_sheet, stat_category, column_index, excel_column, original_column_name,
          cleaned_column_name, role, imported, notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                entry["source_sheet"],
                entry["stat_category"],
                entry["column_index"],
                entry["excel_column"],
                entry["original_column_name"],
                entry["cleaned_column_name"],
                entry["role"],
                1 if entry["imported"] else 0,
                "; ".join(entry["notes"]) if entry["notes"] else None,
            )
            for entry in all_column_entries
        ],
    )
    conn.executemany(
        """
        INSERT INTO stat_values (
          raw_player_name, player_name, player_slug, team, season, season_type, source_sheet, stat_category,
          original_column_name, cleaned_column_name, raw_value, raw_value_json, numeric_value,
          import_notes, source_row_number, source_column_letter
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        all_stat_rows,
    )

    args.profile_dir.mkdir(parents=True, exist_ok=True)
    for stale_profile in args.profile_dir.rglob("*.json"):
        if stale_profile.is_file():
            stale_profile.unlink()

    player_rows: list[tuple[Any, ...]] = []
    players_json: list[dict[str, Any]] = []
    runtime_summaries: list[dict[str, Any]] = []
    for player_slug, player in sorted(players.items(), key=lambda item: item[1]["player_name"].lower()):
        teams_counter: Counter[str] = player["teams"]
        ordered_teams = [team for team, _ in teams_counter.most_common() if team]
        primary_team = ordered_teams[0] if ordered_teams else None
        source_sheets = sorted(player["source_sheets"])
        name_variants = sorted(name for name in player["name_variants"] if name)
        profile_path = f"/data/player_profiles/{player_slug}.json"
        profile = player_profiles[player_slug]
        profile_out = {
            "player_name": profile["player_name"],
            "player_slug": player_slug,
            "season": SEASON,
            "season_type": SEASON_TYPE,
            "teams": sorted(team for team in profile["teams"] if team),
            "primary_team": primary_team,
            "name_variants": sorted(name for name in profile["name_variants"] if name),
            "source_sheets": source_sheets,
            "stat_rows": player["stat_rows"],
            "stats": profile["stats"],
        }
        profile_outputs[player_slug] = profile_out
        write_json(args.profile_dir / f"{player_slug}.json", profile_out, compact=True)
        runtime_summaries.append(build_runtime_summary(profile_out, primary_team))
        player_rows.append(
            (
                player_slug,
                player["player_name"],
                SEASON,
                SEASON_TYPE,
                primary_team,
                json.dumps(ordered_teams, ensure_ascii=False),
                json.dumps(name_variants, ensure_ascii=False),
                json.dumps(source_sheets, ensure_ascii=False),
                player["stat_rows"],
                profile_path,
            )
        )
        players_json.append(
            {
                "player_name": player["player_name"],
                "player_slug": player_slug,
                "season": SEASON,
                "season_type": SEASON_TYPE,
                "primary_team": primary_team,
                "teams": ordered_teams,
                "name_variants": name_variants,
                "source_sheets": source_sheets,
                "stat_rows": player["stat_rows"],
                "profile_path": profile_path,
            }
        )

    conn.executemany(
        """
        INSERT INTO players (
          player_slug, player_name, season, season_type, primary_team, teams_json, name_variants_json,
          source_sheets_json, stat_rows, profile_path
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        player_rows,
    )
    conn.commit()
    conn.close()

    column_dictionary = {
        "generated_at": generated_at,
        "source_workbook": str(workbook_path),
        "source_workbook_sha256": sha256(workbook_path),
        "season": SEASON,
        "season_type": SEASON_TYPE,
        "columns": all_column_entries,
        "by_sheet": {
            sheet_name: [entry for entry in all_column_entries if entry["source_sheet"] == sheet_name]
            for sheet_name in imported_sheets
        },
    }
    write_json(args.column_dictionary, column_dictionary)
    write_json(args.players_json, players_json)
    write_json(args.runtime_summaries, runtime_summaries, compact=True)

    issue_log = {
        "generated_at": generated_at,
        "source_workbook": str(workbook_path),
        "source_workbook_sha256": sha256(workbook_path),
        "season": SEASON,
        "season_type": SEASON_TYPE,
        "audit_summary": audit["summary"],
        "ingestion_summary": {
            "sheets_found": len(audit["sheets"]),
            "sheets_imported": len(imported_sheets),
            "sheets_skipped": len(skipped_sheet_names),
            "sheets_failed": len(failed_sheet_names),
            "unique_players_found": len(players),
            "total_stat_rows_created": len(all_stat_rows),
            "column_dictionary_entries": len(all_column_entries),
            "issues_logged": len(issues),
        },
        "sheets_imported": imported_sheets,
        "sheets_skipped": skipped_sheet_names,
        "sheets_failed": failed_sheet_names,
        "issues": issues,
    }
    write_json(args.issues_log, issue_log)

    postgres_summary = {"requested": False, "written": False}
    if args.write_postgres:
        postgres_summary = write_postgres_outputs(
            args=args,
            generated_at=generated_at,
            workbook_path=workbook_path,
            audit=audit,
            imported_sheets=imported_sheets,
            skipped_sheet_names=skipped_sheet_names,
            failed_sheet_names=failed_sheet_names,
            all_column_entries=all_column_entries,
            all_stat_rows=all_stat_rows,
            players_json=players_json,
            profile_outputs=profile_outputs,
            runtime_summaries=runtime_summaries,
            issue_log=issue_log,
        )

    example_slug = None
    for preferred in ["Nikola Jokić", "Luka Dončić", "Shai Gilgeous-Alexander", "Giannis Antetokounmpo"]:
        preferred_slug = slugify(preferred)
        if preferred_slug in player_profiles:
            example_slug = preferred_slug
            break
    if example_slug is None and players_json:
        example_slug = players_json[0]["player_slug"]

    return {
        "sheets_found": len(audit["sheets"]),
        "sheets_imported": imported_sheets,
        "sheets_skipped": skipped_sheet_names,
        "sheets_failed": failed_sheet_names,
        "unique_players_found": len(players),
        "total_stat_rows_created": len(all_stat_rows),
        "top_20_issues": issues[:20],
        "example_profile_path": str(args.profile_dir / f"{example_slug}.json") if example_slug else None,
        "outputs": {
            "sqlite": str(args.sqlite),
            "column_dictionary": str(args.column_dictionary),
            "issues_log": str(args.issues_log),
            "players_json": str(args.players_json),
            "profile_dir": str(args.profile_dir),
            "runtime_summaries": str(args.runtime_summaries),
        },
        "postgres": postgres_summary,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest clean NBA Excel sheets into normalized outputs.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sqlite", type=Path, default=DEFAULT_SQLITE)
    parser.add_argument("--column-dictionary", type=Path, default=DEFAULT_COLUMN_DICTIONARY)
    parser.add_argument("--issues-log", type=Path, default=DEFAULT_ISSUES_LOG)
    parser.add_argument("--players-json", type=Path, default=DEFAULT_PLAYERS_JSON)
    parser.add_argument("--profile-dir", type=Path, default=DEFAULT_PROFILE_DIR)
    parser.add_argument("--runtime-summaries", type=Path, default=DEFAULT_RUNTIME_SUMMARIES)
    parser.add_argument("--official-snapshot", type=Path, default=DEFAULT_OFFICIAL_SNAPSHOT)
    parser.add_argument("--write-postgres", action="store_true", help="Write processed output to Postgres after JSON/SQLite generation.")
    parser.add_argument("--postgres-batch-size", type=int, default=5000)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.write_postgres and not database_url_from_env():
        print(
            "Postgres write requested with --write-postgres, but DATABASE_URL is not set. "
            "Run without --write-postgres for JSON/SQLite generation only, or set DATABASE_URL.",
            file=sys.stderr,
        )
        raise SystemExit(2)
    summary = import_workbook(args)
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
