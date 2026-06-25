#!/usr/bin/env python3
"""Validate the master workbook and refresh Postgres only when required."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "raw" / "nba_data_2025_26.xlsx"
MINIMUM_SHEETS = 60
MINIMUM_PLAYERS = 500
MINIMUM_STAT_ROWS = 500_000
EXPECTED_TEAMS = 30
MINIMUM_GAMES = 1_300
MINIMUM_TEAM_GAME_STATS = 2_600
MINIMUM_PLAYER_GAME_STATS = 28_000


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def database_url() -> str | None:
    value = os.environ.get("DATABASE_URL", "").strip()
    return value or None


def validate_workbook(workbook_path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to validate the master workbook. Install requirements.txt.") from exc

    if not workbook_path.exists():
        raise RuntimeError(f"Master workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    if len(sheet_names) < MINIMUM_SHEETS:
        raise RuntimeError(
            f"Master workbook validation failed: found {len(sheet_names)} sheets; expected at least {MINIMUM_SHEETS}."
        )
    return {
        "path": str(workbook_path),
        "sha256": sha256(workbook_path),
        "size_bytes": workbook_path.stat().st_size,
        "sheets": len(sheet_names),
        "sheet_names": sheet_names,
    }


def probe_database(connection_string: str) -> dict[str, Any]:
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for production refreshes. Install requirements.txt.") from exc

    with psycopg.connect(connection_string) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                  to_regclass('public.current_ingestion_run') IS NOT NULL,
                  to_regclass('public.current_player_season_summaries') IS NOT NULL,
                  to_regclass('public.current_team_season_summaries') IS NOT NULL,
                  to_regclass('public.current_games') IS NOT NULL,
                  to_regclass('public.current_team_game_stats') IS NOT NULL,
                  to_regclass('public.current_player_game_stats') IS NOT NULL
                """
            )
            schema_flags = cursor.fetchone()
            if not schema_flags or not all(schema_flags):
                raise RuntimeError("Postgres schema is incomplete. Run `pnpm db:migrate` before refreshing data.")

            cursor.execute(
                """
                SELECT
                  run.id::text,
                  run.source_workbook_sha256,
                  run.finished_at,
                  run.unique_players,
                  run.stat_rows_created,
                  (SELECT count(*) FROM current_player_season_summaries),
                  (SELECT count(*) FROM current_team_season_summaries),
                  (SELECT count(*) FROM current_games),
                  (SELECT count(*) FROM current_team_game_stats),
                  (SELECT count(*) FROM current_player_game_stats)
                FROM current_ingestion_run run
                UNION ALL
                SELECT NULL, NULL, NULL, 0, 0, 0, 0, 0, 0, 0
                WHERE NOT EXISTS (SELECT 1 FROM current_ingestion_run)
                LIMIT 1
                """
            )
            row = cursor.fetchone()
    if not row:
        raise RuntimeError("Unable to read current ingestion state.")
    return {
        "ingestion_run_id": row[0],
        "source_workbook_sha256": row[1],
        "finished_at": row[2].isoformat() if row[2] else None,
        "players_reported": int(row[3] or 0),
        "stat_rows_reported": int(row[4] or 0),
        "player_summaries": int(row[5] or 0),
        "team_summaries": int(row[6] or 0),
        "games": int(row[7] or 0),
        "team_game_stats": int(row[8] or 0),
        "player_game_stats": int(row[9] or 0),
    }


def database_is_complete(state: dict[str, Any]) -> bool:
    return (
        state["players_reported"] >= MINIMUM_PLAYERS
        and state["stat_rows_reported"] >= MINIMUM_STAT_ROWS
        and state["player_summaries"] >= MINIMUM_PLAYERS
        and state["team_summaries"] == EXPECTED_TEAMS
        and state["games"] >= MINIMUM_GAMES
        and state["team_game_stats"] >= MINIMUM_TEAM_GAME_STATS
        and state["player_game_stats"] >= MINIMUM_PLAYER_GAME_STATS
    )


def run_ingestion(workbook_path: Path, connection_string: str) -> dict[str, Any]:
    with tempfile.TemporaryDirectory(prefix="basketball-savant-refresh-") as temp_directory:
        output = Path(temp_directory)
        profile_directory = output / "profiles"
        profile_directory.mkdir()
        command = [
            sys.executable,
            str(ROOT / "scripts" / "ingest_nba_excel.py"),
            "--workbook",
            str(workbook_path),
            "--write-postgres",
            "--postgres-batch-size",
            "10000",
            "--sqlite",
            str(output / "master.sqlite"),
            "--column-dictionary",
            str(output / "columns.json"),
            "--issues-log",
            str(output / "issues.json"),
            "--players-json",
            str(output / "players.json"),
            "--profile-dir",
            str(profile_directory),
            "--runtime-summaries",
            str(output / "summaries.json"),
            "--runtime-fallbacks",
            str(output / "runtime-fallbacks.json"),
        ]
        environment = {**os.environ, "DATABASE_URL": connection_string}
        completed = subprocess.run(command, cwd=ROOT, env=environment, text=True, capture_output=True)
        if completed.returncode != 0:
            detail = completed.stderr.strip() or completed.stdout.strip() or "unknown ingestion error"
            raise RuntimeError(f"Masterfile ingestion failed: {detail}")
        return json.loads(completed.stdout)


def prune_ingestion_history(connection_string: str, retain_runs: int) -> int:
    if retain_runs < 1:
        return 0
    import psycopg

    with psycopg.connect(connection_string) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                WITH old_runs AS (
                  SELECT id
                  FROM ingestion_runs
                  WHERE status = 'succeeded'
                  ORDER BY finished_at DESC NULLS LAST, started_at DESC
                  OFFSET %s
                )
                DELETE FROM ingestion_runs
                WHERE id IN (SELECT id FROM old_runs)
                """,
                (retain_runs,),
            )
            return cursor.rowcount


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--write-postgres", action="store_true")
    parser.add_argument("--force", action="store_true", help="Write a new ingestion even when the workbook checksum is unchanged.")
    parser.add_argument("--retain-runs", type=int, default=2)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    connection_string = database_url()

    if args.write_postgres and not connection_string:
        print("DATABASE_URL is required when --write-postgres is enabled.", file=sys.stderr)
        raise SystemExit(2)

    workbook = validate_workbook(args.workbook.resolve())
    if not args.write_postgres:
        print(json.dumps({"status": "validated", "workbook": workbook}, indent=2))
        return

    before = probe_database(connection_string)
    should_write = (
        args.force
        or before["source_workbook_sha256"] != workbook["sha256"]
        or not database_is_complete(before)
    )
    if not should_write:
        print(json.dumps({
            "status": "skipped_unchanged",
            "workbook": workbook,
            "database": before,
        }, indent=2))
        return

    ingestion = run_ingestion(args.workbook.resolve(), connection_string)
    after = probe_database(connection_string)
    if after["source_workbook_sha256"] != workbook["sha256"] or not database_is_complete(after):
        raise RuntimeError("Post-ingestion validation failed; current database views are incomplete.")
    pruned_runs = prune_ingestion_history(connection_string, args.retain_runs)

    print(json.dumps({
        "status": "refreshed",
        "workbook": workbook,
        "database_before": before,
        "database_after": after,
        "ingestion": ingestion["postgres"],
        "pruned_runs": pruned_runs,
    }, indent=2))


if __name__ == "__main__":
    main()
