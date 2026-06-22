#!/usr/bin/env python3
"""Audit the raw NBA master workbook without modifying it."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "raw" / "nba_data_2025_26.xlsx"
DEFAULT_AUDIT_MD = ROOT / "data" / "processed" / "nba_excel_import_audit.md"
DEFAULT_ISSUES_JSON = ROOT / "data" / "processed" / "data_issues_log.json"
ORIGINAL_UPLOAD = Path("/Users/johnnypark/Downloads/nba_data_2025_26.xlsx")

COMMON_HEADER_TOKENS = {
    "age",
    "ast",
    "ast%",
    "blk",
    "blka",
    "defrtg",
    "dfg%",
    "dreb",
    "efg%",
    "fg%",
    "fga",
    "fgm",
    "ft%",
    "fta",
    "ftm",
    "g",
    "gp",
    "min",
    "netrtg",
    "oreb",
    "pace",
    "pf",
    "pie",
    "pts",
    "reb",
    "stl",
    "team",
    "tov",
    "usg%",
    "w",
    "l",
}

CONTROL_ROW_TOKENS = {
    "conference",
    "date from",
    "date to",
    "distance range",
    "division",
    "draft pick",
    "draft year",
    "group quantity",
    "league",
    "month",
    "outcome",
    "pace adjust",
    "per mode",
    "period",
    "player experience",
    "player position",
    "players",
    "season",
    "season segment",
    "season type",
    "shot clock range",
    "sort by",
    "starter bench",
    "vs conference",
    "vs division",
}

NBA_EXPORT_MARKERS = CONTROL_ROW_TOKENS | {
    "nba",
    "all teams",
    "all players",
    "regular season",
    "playoffs",
}

POSITION_ORDER = {"PG": 1, "SG": 2, "SF": 3, "PF": 4, "C": 5}


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize(value: Any) -> str:
    value_text = text_value(value)
    value_text = value_text.replace("\n", " ").replace("\r", " ")
    value_text = re.sub(r"\s+", " ", value_text).strip().lower()
    return value_text


def normalized_header_label(value: Any) -> str:
    label = normalize(value)
    label = label.replace(".", "")
    return label


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def row_values(ws, row_idx: int) -> list[Any]:
    return [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]


def row_texts(ws, row_idx: int) -> list[str]:
    return [text_value(value) for value in row_values(ws, row_idx)]


def compact_list(values: list[Any], limit: int = 20) -> list[Any]:
    if len(values) <= limit:
        return values
    return values[:limit] + [f"... +{len(values) - limit} more"]


def compact_ranges(indices: list[int], limit: int = 12) -> list[str]:
    if not indices:
        return []
    ranges: list[str] = []
    start = prev = indices[0]
    for idx in indices[1:]:
        if idx == prev + 1:
            prev = idx
            continue
        ranges.append(str(start) if start == prev else f"{start}-{prev}")
        start = prev = idx
    ranges.append(str(start) if start == prev else f"{start}-{prev}")
    if len(ranges) <= limit:
        return ranges
    return ranges[:limit] + [f"... +{len(ranges) - limit} more ranges"]


def infer_category(sheet_name: str) -> str:
    lowered = sheet_name.lower()
    if lowered == "general - official leaders":
        return "Official leaders"
    if lowered.startswith("general -"):
        return "General player stats"
    if lowered.startswith("clutch -"):
        return "Clutch player stats"
    if lowered.startswith("playtype -"):
        return "Play type"
    if re.fullmatch(r"sheet\d+", sheet_name, flags=re.IGNORECASE):
        return "Unnamed NBA.com play type export"
    if lowered.startswith("tracking -"):
        return "Tracking"
    if lowered.startswith("defense dashboard -"):
        return "Defense dashboard"
    if lowered.startswith("shooting dashboard -"):
        return "Shooting dashboard"
    if lowered.startswith("opponent shooting -"):
        return "Opponent shooting zones/ranges"
    if lowered.startswith("shooting -"):
        return "Player shooting zones/ranges"
    if lowered == "hustle":
        return "Hustle"
    if lowered == "box outs":
        return "Box outs"
    if lowered == "bios":
        return "Player bios and physical profile"
    return "Unknown"


def header_score(labels: list[str]) -> float:
    nonblank = [label for label in labels if label]
    if len(nonblank) < 4:
        return -1
    label_set = set(nonblank)
    exact_player = "player" in label_set or any(label.endswith(" player") for label in label_set)
    exact_team = "team" in label_set
    common_count = sum(1 for label in nonblank if label in COMMON_HEADER_TOKENS)
    score = len(nonblank) * 0.5 + common_count * 4
    if exact_player:
        score += 20
    if exact_team:
        score += 16
    if {"player", "team", "gp"}.issubset(label_set):
        score += 15
    if CONTROL_ROW_TOKENS.intersection(label_set):
        score -= 18
    if not exact_player and not exact_team:
        score -= 10
    return score


def detect_header_row(ws) -> tuple[int | None, float]:
    best_row = None
    best_score = -1.0
    for row_idx in range(1, min(ws.max_row, 90) + 1):
        labels = [normalized_header_label(value) for value in row_values(ws, row_idx)]
        score = header_score(labels)
        if score > best_score:
            best_score = score
            best_row = row_idx
    if best_score < 20:
        return None, best_score
    return best_row, best_score


def duplicate_headers(headers: list[str]) -> list[dict[str, Any]]:
    normalized_to_columns: dict[str, list[str]] = defaultdict(list)
    original_by_normalized: dict[str, str] = {}
    for idx, header in enumerate(headers, start=1):
        normalized = normalized_header_label(header)
        if not normalized:
            continue
        normalized_to_columns[normalized].append(get_column_letter(idx))
        original_by_normalized.setdefault(normalized, header)
    duplicates = []
    for normalized, columns in normalized_to_columns.items():
        if len(columns) > 1:
            duplicates.append(
                {
                    "name": original_by_normalized[normalized],
                    "normalizedName": normalized,
                    "columns": columns,
                }
            )
    return duplicates


def detect_repeated_header_rows(ws, header_row: int | None, headers: list[str]) -> list[int]:
    if header_row is None:
        return []
    header_norm = [normalized_header_label(value) for value in headers]
    header_nonblank_positions = [idx for idx, label in enumerate(header_norm) if label]
    repeated: list[int] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        labels = [normalized_header_label(value) for value in row_values(ws, row_idx)]
        matches = 0
        compared = 0
        for idx in header_nonblank_positions:
            if idx >= len(labels):
                continue
            compared += 1
            if labels[idx] == header_norm[idx]:
                matches += 1
        if compared >= 4 and matches / compared >= 0.7:
            repeated.append(row_idx)
    return repeated


def looks_like_nba_export(ws, header_row: int | None) -> bool:
    search_limit = min(ws.max_row, max(header_row or 0, 30))
    marker_hits = 0
    for row_idx in range(1, search_limit + 1):
        for value in row_values(ws, row_idx):
            normalized = normalize(value)
            if normalized in NBA_EXPORT_MARKERS:
                marker_hits += 1
            elif any(marker in normalized for marker in ("regular season", "all teams", "all players")):
                marker_hits += 1
    return marker_hits >= 3


def detect_grouped_header(ws, header_row: int | None, duplicates: list[dict[str, Any]]) -> bool:
    if header_row is None or header_row <= 1:
        return False
    previous = [normalize(value) for value in row_values(ws, header_row - 1)]
    previous_nonblank = [value for value in previous if value]
    zone_words = {"restricted area", "mid-range", "corner", "paint", "pullups", "catch and shoot"}
    has_zone_label = any(any(word in value for word in zone_words) for value in previous_nonblank)
    repeated_metric_names = sum(1 for item in duplicates if item["normalizedName"] in {"fgm", "fga", "fg%", "3pm", "3pa", "3p%"})
    return bool(previous_nonblank) and (has_zone_label or repeated_metric_names >= 2)


def detect_height_serials(ws, header_row: int | None, headers: list[str]) -> dict[str, Any] | None:
    if header_row is None:
        return None
    height_col = None
    for idx, header in enumerate(headers, start=1):
        if normalized_header_label(header) == "height":
            height_col = idx
            break
    if height_col is None:
        return None
    suspicious_rows: list[dict[str, Any]] = []
    checked = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=height_col).value
        if value in (None, ""):
            continue
        checked += 1
        if isinstance(value, (int, float)) and value > 100:
            player = text_value(ws.cell(row=row_idx, column=1).value)
            suspicious_rows.append({"row": row_idx, "player": player, "value": value})
        elif isinstance(value, (datetime, date, time)):
            player = text_value(ws.cell(row=row_idx, column=1).value)
            suspicious_rows.append({"row": row_idx, "player": player, "value": text_value(value)})
    if suspicious_rows:
        return {
            "checkedRows": checked,
            "suspiciousCount": len(suspicious_rows),
            "examples": compact_list(suspicious_rows, 10),
        }
    return None


def import_decision(
    *,
    header_row: int | None,
    has_player: bool,
    has_team: bool,
    data_row_count: int,
    issues: list[dict[str, Any]],
) -> str:
    if header_row is None or data_row_count == 0:
        return "skip"
    if not has_player or not has_team:
        return "review"
    if any(issue["severity"] == "major" for issue in issues):
        return "review"
    return "import"


def issue(severity: str, issue_type: str, message: str, **details: Any) -> dict[str, Any]:
    return {
        "severity": severity,
        "type": issue_type,
        "message": message,
        "details": details,
    }


def audit_sheet(ws) -> dict[str, Any]:
    header_row, score = detect_header_row(ws)
    headers = row_texts(ws, header_row) if header_row else []
    header_labels = [normalized_header_label(header) for header in headers]
    label_set = set(label for label in header_labels if label)
    has_player = "player" in label_set or any(label.endswith(" player") for label in label_set)
    has_team = "team" in label_set
    data_row_count = 0
    if header_row is not None:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if any(text_value(value) for value in row_values(ws, row_idx)):
                data_row_count += 1

    used_rows = [
        row_idx
        for row_idx in range(1, ws.max_row + 1)
        if any(text_value(value) for value in row_values(ws, row_idx))
    ]
    used_cols = []
    for col_idx in range(1, ws.max_column + 1):
        if any(text_value(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(1, ws.max_row + 1)):
            used_cols.append(col_idx)

    blank_rows = [row_idx for row_idx in range(1, ws.max_row + 1) if row_idx not in set(used_rows)]
    blank_cols = [col_idx for col_idx in range(1, ws.max_column + 1) if col_idx not in set(used_cols)]
    blank_header_cols = []
    if header_row is not None:
        for col_idx in used_cols:
            if not text_value(ws.cell(row=header_row, column=col_idx).value):
                blank_header_cols.append(get_column_letter(col_idx))

    duplicate_columns = duplicate_headers(headers)
    repeated_headers = detect_repeated_header_rows(ws, header_row, headers)
    grouped_header = detect_grouped_header(ws, header_row, duplicate_columns)
    header_merges = []
    all_merges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
    for merged_range in ws.merged_cells.ranges:
        if header_row is not None and merged_range.min_row <= header_row <= merged_range.max_row:
            header_merges.append(str(merged_range))
        elif header_row is not None and merged_range.min_row <= max(header_row - 1, 1) <= merged_range.max_row:
            header_merges.append(str(merged_range))

    non_text_headers = []
    if header_row is not None:
        for col_idx in used_cols:
            value = ws.cell(row=header_row, column=col_idx).value
            if isinstance(value, (int, float, datetime, date, time)):
                non_text_headers.append({"column": get_column_letter(col_idx), "value": text_value(value)})

    height_serials = detect_height_serials(ws, header_row, headers)
    category = infer_category(ws.title)
    nba_export = looks_like_nba_export(ws, header_row)
    helper_like = not nba_export and data_row_count < 10

    issues: list[dict[str, Any]] = []
    if header_row is None:
        issues.append(issue("major", "missing_header", "No reliable table header row was detected."))
    if not has_player:
        issues.append(issue("major", "missing_player_column", "No Player column was detected on the likely header row."))
    if not has_team:
        issues.append(issue("major", "missing_team_column", "No Team column was detected on the likely header row."))
    if data_row_count == 0:
        issues.append(issue("major", "no_data_rows", "No nonblank data rows were detected after the likely header."))
    if re.fullmatch(r"sheet\d+", ws.title, flags=re.IGNORECASE):
        issues.append(
            issue(
                "major",
                "generic_sheet_name",
                "Sheet uses a generic Excel name and should be identified before import.",
                sheetName=ws.title,
                inferredCategory=category,
            )
        )
    if grouped_header:
        issues.append(
            issue(
                "major",
                "multi_row_grouped_header",
                "Header appears to use grouped NBA.com zone/range labels and must be flattened before import.",
                previousRow=header_row - 1 if header_row else None,
            )
        )
    if non_text_headers:
        issues.append(
            issue(
                "major",
                "non_text_header_values",
                "Likely header row contains date/time/numeric values where metric names were expected.",
                examples=compact_list(non_text_headers, 10),
            )
        )
    if height_serials:
        issues.append(
            issue(
                "major",
                "height_serial_or_date_values",
                "Height column includes Excel date/serial-like values that need normalization.",
                **height_serials,
            )
        )
    if duplicate_columns:
        issues.append(
            issue(
                "warning",
                "duplicate_column_names",
                "Duplicate column labels were found on the likely header row.",
                duplicates=compact_list(duplicate_columns, 20),
            )
        )
    if blank_header_cols:
        issues.append(
            issue(
                "warning",
                "blank_header_columns",
                "One or more used columns have a blank header cell.",
                columns=compact_list(blank_header_cols, 20),
            )
        )
    if blank_cols:
        issues.append(
            issue(
                "info",
                "blank_columns",
                "Blank columns exist within the worksheet's used range.",
                count=len(blank_cols),
                columns=compact_list([get_column_letter(idx) for idx in blank_cols], 20),
            )
        )
    if blank_rows:
        issues.append(
            issue(
                "info",
                "blank_rows",
                "Blank rows exist within the worksheet's used range.",
                count=len(blank_rows),
                rowRanges=compact_ranges(blank_rows),
            )
        )
    if repeated_headers:
        issues.append(
            issue(
                "warning",
                "repeated_header_rows",
                "Likely table header is repeated inside the data area.",
                rows=compact_list(repeated_headers, 20),
            )
        )
    if header_merges:
        issues.append(
            issue(
                "warning",
                "merged_header_cells",
                "Merged cells touch the likely header or grouped header rows.",
                ranges=compact_list(header_merges, 20),
            )
        )
    elif all_merges:
        issues.append(
            issue(
                "info",
                "merged_cells",
                "Merged cells exist elsewhere in the worksheet.",
                count=len(all_merges),
                ranges=compact_list(all_merges, 10),
            )
        )
    if not nba_export:
        issues.append(
            issue(
                "warning",
                "not_obvious_nba_export",
                "Sheet does not contain enough NBA.com export markers in the top rows.",
            )
        )

    decision = import_decision(
        header_row=header_row,
        has_player=has_player,
        has_team=has_team,
        data_row_count=data_row_count,
        issues=issues,
    )

    return {
        "sheetName": ws.title,
        "rowCount": ws.max_row,
        "columnCount": ws.max_column,
        "nonblankRowCount": len(used_rows),
        "nonblankColumnCount": len(used_cols),
        "likelyHeaderRow": header_row,
        "headerDetectionScore": round(score, 2),
        "headerPreview": headers,
        "hasPlayerColumn": has_player,
        "hasTeamColumn": has_team,
        "dataRowCountAfterHeader": data_row_count,
        "blankRows": {"count": len(blank_rows), "ranges": compact_ranges(blank_rows)},
        "blankColumns": {
            "count": len(blank_cols),
            "columns": compact_list([get_column_letter(idx) for idx in blank_cols], 20),
        },
        "blankHeaderColumns": blank_header_cols,
        "duplicateColumnNames": duplicate_columns,
        "repeatedHeaderRows": repeated_headers,
        "mergedCells": {"count": len(all_merges), "ranges": compact_list(all_merges, 20)},
        "headerMergedCells": header_merges,
        "groupedOrMultiRowHeader": grouped_header,
        "nbaDotComExportLikely": nba_export,
        "manualNotesOrHelperLikely": helper_like,
        "possibleStatCategory": category,
        "decision": decision,
        "issues": issues,
    }


def md_escape(value: Any) -> str:
    text = text_value(value)
    return text.replace("|", "\\|").replace("\n", " ")


def short_issue_list(sheet: dict[str, Any]) -> str:
    if not sheet["issues"]:
        return "None"
    major = [item["type"] for item in sheet["issues"] if item["severity"] == "major"]
    warnings = [item["type"] for item in sheet["issues"] if item["severity"] == "warning"]
    ordered = major + warnings
    if not ordered:
        ordered = [item["type"] for item in sheet["issues"][:2]]
    return ", ".join(ordered[:5])


def render_markdown(audit: dict[str, Any]) -> str:
    workbook = audit["workbook"]
    summary = audit["summary"]
    lines: list[str] = []
    lines.append("# NBA Excel Import Audit")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Exact file path used: `{workbook['path']}`")
    lines.append(f"- Workbook name: `{workbook['name']}`")
    lines.append(f"- SHA256: `{workbook['sha256']}`")
    lines.append(f"- Original upload path checked: `{workbook['originalUploadPath']}`")
    lines.append(f"- Original upload checksum matches raw copy: `{workbook['originalUploadChecksumMatches']}`")
    lines.append(f"- Sheets found: {summary['sheetsFound']}")
    lines.append(f"- Importable sheets: {summary['importableSheets']}")
    lines.append(f"- Needs review: {summary['needsReviewSheets']}")
    lines.append(f"- Skipped sheets: {summary['skippedSheets']}")
    lines.append(f"- Major issues: {summary['majorIssues']}")
    lines.append("")
    lines.append("## Sheet Names")
    lines.append("")
    for idx, name in enumerate(workbook["sheetNames"], start=1):
        lines.append(f"{idx}. `{name}`")
    lines.append("")
    lines.append("## Workbook-Wide Findings")
    lines.append("")
    lines.append("- The workbook appears to be primarily NBA.com export sheets, with filter/control rows above the usable table headers.")
    lines.append("- Most usable table headers begin around rows 18-20, not row 1.")
    lines.append("- The raw workbook should remain immutable; import code should write normalized data to `data/processed/` or `public/data/` only.")
    if summary["genericSheetNames"]:
        lines.append(f"- Generic sheet names requiring review: {', '.join(f'`{name}`' for name in summary['genericSheetNames'])}.")
    if summary["multiRowHeaderSheets"]:
        lines.append(f"- Multi-row/grouped header sheets requiring flattening: {', '.join(f'`{name}`' for name in summary['multiRowHeaderSheets'])}.")
    if summary["nonTextHeaderSheets"]:
        lines.append(f"- Sheets with date/time/numeric values in the detected header: {', '.join(f'`{name}`' for name in summary['nonTextHeaderSheets'])}.")
    if summary["heightSerialSheets"]:
        lines.append(f"- Sheets with Excel date/serial-like height values: {', '.join(f'`{name}`' for name in summary['heightSerialSheets'])}.")
    if summary["helperLikeSheets"]:
        lines.append(f"- Sheets that look like manual notes/helper sheets: {', '.join(f'`{name}`' for name in summary['helperLikeSheets'])}.")
    else:
        lines.append("- Sheets that look like manual notes/helper sheets: none detected.")
    lines.append("")
    lines.append("## Decision Buckets")
    lines.append("")
    lines.append(f"- Clean import candidates ({len(summary['importSheetNames'])}): {', '.join(f'`{name}`' for name in summary['importSheetNames'])}")
    lines.append(f"- Needs manual review before import ({len(summary['reviewSheetNames'])}): {', '.join(f'`{name}`' for name in summary['reviewSheetNames'])}")
    lines.append(f"- Skip candidates ({len(summary['skippedSheetNames'])}): {', '.join(f'`{name}`' for name in summary['skippedSheetNames']) if summary['skippedSheetNames'] else 'none'}")
    lines.append("")
    lines.append("## Sheet Inventory")
    lines.append("")
    lines.append("| # | Sheet | Rows | Cols | Header Row | Player Col | Team Col | Category | Decision | Key Issues |")
    lines.append("|---:|---|---:|---:|---:|:---:|:---:|---|---|---|")
    for idx, sheet in enumerate(audit["sheets"], start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(idx),
                    f"`{md_escape(sheet['sheetName'])}`",
                    str(sheet["rowCount"]),
                    str(sheet["columnCount"]),
                    str(sheet["likelyHeaderRow"] or "n/a"),
                    "yes" if sheet["hasPlayerColumn"] else "no",
                    "yes" if sheet["hasTeamColumn"] else "no",
                    md_escape(sheet["possibleStatCategory"]),
                    sheet["decision"],
                    md_escape(short_issue_list(sheet)),
                ]
            )
            + " |"
        )
    lines.append("")
    lines.append("## Per-Sheet Audit Details")
    for sheet in audit["sheets"]:
        lines.append("")
        lines.append(f"### {sheet['sheetName']}")
        lines.append("")
        lines.append(f"- Rows x columns: {sheet['rowCount']} x {sheet['columnCount']}")
        lines.append(f"- Nonblank rows x columns: {sheet['nonblankRowCount']} x {sheet['nonblankColumnCount']}")
        lines.append(f"- Likely header row: {sheet['likelyHeaderRow'] or 'not detected'}")
        lines.append(f"- Player/team columns: player={sheet['hasPlayerColumn']}, team={sheet['hasTeamColumn']}")
        lines.append(f"- NBA.com export likely: {sheet['nbaDotComExportLikely']}")
        lines.append(f"- Manual notes/helper likely: {sheet['manualNotesOrHelperLikely']}")
        lines.append(f"- Possible stat category: {sheet['possibleStatCategory']}")
        lines.append(f"- Import decision: {sheet['decision']}")
        lines.append(f"- Blank rows: {sheet['blankRows']['count']} ({', '.join(sheet['blankRows']['ranges']) if sheet['blankRows']['ranges'] else 'none'})")
        lines.append(
            f"- Blank columns: {sheet['blankColumns']['count']} ({', '.join(sheet['blankColumns']['columns']) if sheet['blankColumns']['columns'] else 'none'})"
        )
        if sheet["blankHeaderColumns"]:
            lines.append(f"- Blank header columns: {', '.join(sheet['blankHeaderColumns'])}")
        else:
            lines.append("- Blank header columns: none")
        if sheet["duplicateColumnNames"]:
            dupes = [
                f"{item['name']} ({', '.join(item['columns'])})"
                for item in sheet["duplicateColumnNames"]
            ]
            lines.append(f"- Duplicate column names: {', '.join(dupes)}")
        else:
            lines.append("- Duplicate column names: none")
        if sheet["repeatedHeaderRows"]:
            lines.append(f"- Repeated header rows inside data: {', '.join(map(str, sheet['repeatedHeaderRows']))}")
        else:
            lines.append("- Repeated header rows inside data: none")
        if sheet["groupedOrMultiRowHeader"]:
            lines.append("- Weird merged/header formatting: grouped or multi-row header detected")
        elif sheet["headerMergedCells"]:
            lines.append(f"- Weird merged/header formatting: merged header cells {', '.join(sheet['headerMergedCells'])}")
        elif sheet["mergedCells"]["count"]:
            lines.append(f"- Weird merged/header formatting: {sheet['mergedCells']['count']} merged cells outside likely header")
        else:
            lines.append("- Weird merged/header formatting: none detected")
        if sheet["issues"]:
            lines.append("- Issues:")
            for item in sheet["issues"]:
                lines.append(f"  - [{item['severity']}] {item['type']}: {item['message']}")
        else:
            lines.append("- Issues: none")
    lines.append("")
    lines.append("## Safest Import Plan")
    lines.append("")
    lines.append("1. Keep `data/raw/nba_data_2025_26.xlsx` as the immutable raw source of truth and never import directly from frontend code.")
    lines.append("2. Build a read-only importer that normalizes each approved sheet into typed JSON/CSV outputs under `data/processed/`, then publish only validated frontend-ready data under `public/data/`.")
    lines.append("3. Phase 1 should import clean single-header player tables marked `import`, keyed by normalized player name, NBA Stats player ID when available, team abbreviation, season, and stat category.")
    lines.append("4. Phase 2 should handle `review` sheets with duplicate/grouped headers by explicitly flattening parent header rows into stable names, especially shooting zones and opponent shooting zones.")
    lines.append("5. Phase 3 should normalize `Bios` after resolving Excel serial-like height values, then use it for height, weight, college, country, draft fields, and profile metadata.")
    lines.append("6. Before using any generated data in the app, add validation for duplicate players, missing teams, nonnumeric stat cells, null sorting behavior, and row-count drift against this audit.")
    lines.append("7. Defer any skipped or unresolved sheets until their source layout is mapped manually; do not hardcode fixes in React components.")
    lines.append("")
    return "\n".join(lines)


def build_audit(workbook_path: Path) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    sheets = [audit_sheet(ws) for ws in wb.worksheets]
    issue_entries = []
    for sheet in sheets:
        for item in sheet["issues"]:
            issue_entries.append({"sheetName": sheet["sheetName"], **item})

    decision_counts = Counter(sheet["decision"] for sheet in sheets)
    major_issues = [entry for entry in issue_entries if entry["severity"] == "major"]
    original_hash = sha256(ORIGINAL_UPLOAD) if ORIGINAL_UPLOAD.exists() else None
    raw_hash = sha256(workbook_path)
    summary = {
        "sheetsFound": len(sheets),
        "importableSheets": decision_counts.get("import", 0),
        "needsReviewSheets": decision_counts.get("review", 0),
        "skippedSheets": decision_counts.get("skip", 0),
        "majorIssues": len(major_issues),
        "genericSheetNames": [sheet["sheetName"] for sheet in sheets if any(item["type"] == "generic_sheet_name" for item in sheet["issues"])],
        "multiRowHeaderSheets": [sheet["sheetName"] for sheet in sheets if sheet["groupedOrMultiRowHeader"]],
        "nonTextHeaderSheets": [sheet["sheetName"] for sheet in sheets if any(item["type"] == "non_text_header_values" for item in sheet["issues"])],
        "heightSerialSheets": [sheet["sheetName"] for sheet in sheets if any(item["type"] == "height_serial_or_date_values" for item in sheet["issues"])],
        "helperLikeSheets": [sheet["sheetName"] for sheet in sheets if sheet["manualNotesOrHelperLikely"]],
        "importSheetNames": [sheet["sheetName"] for sheet in sheets if sheet["decision"] == "import"],
        "reviewSheetNames": [sheet["sheetName"] for sheet in sheets if sheet["decision"] == "review"],
        "skippedSheetNames": [sheet["sheetName"] for sheet in sheets if sheet["decision"] == "skip"],
    }
    return {
        "workbook": {
            "path": str(workbook_path.resolve()),
            "name": workbook_path.name,
            "sha256": raw_hash,
            "sheetCount": len(sheets),
            "sheetNames": [sheet["sheetName"] for sheet in sheets],
            "originalUploadPath": str(ORIGINAL_UPLOAD) if ORIGINAL_UPLOAD.exists() else None,
            "originalUploadSha256": original_hash,
            "originalUploadChecksumMatches": bool(original_hash and original_hash == raw_hash),
        },
        "summary": summary,
        "issues": issue_entries,
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit the NBA master workbook without modifying it.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--audit-md", type=Path, default=DEFAULT_AUDIT_MD)
    parser.add_argument("--issues-json", type=Path, default=DEFAULT_ISSUES_JSON)
    args = parser.parse_args()

    args.audit_md.parent.mkdir(parents=True, exist_ok=True)
    args.issues_json.parent.mkdir(parents=True, exist_ok=True)
    audit = build_audit(args.workbook)
    args.audit_md.write_text(render_markdown(audit), encoding="utf-8")
    args.issues_json.write_text(json.dumps(audit, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(audit["summary"], indent=2))


if __name__ == "__main__":
    main()
